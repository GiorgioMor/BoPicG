import json
import os
from openpyxl import load_workbook

wb = load_workbook('F:/Giorgio/Multimedia/Download/proverbi.xlsx')
ws = wb.active

num_cells_x = int((ws.max_row))

cards = []

for x in range(1, num_cells_x):
    card_cell = ws.cell(x,1)
    card = {}
    print(card_cell.value.upper())
    card = {
        'cat': 'S',
        'num': 1,
        'par': card_cell.value.upper(),
    }
    cards.append(card)

j = json.dumps(cards, ensure_ascii=False)
print(j)

with open(os.path.join(os.path.dirname(os.path.realpath(__file__)), 'more_cards.json'), 'w', encoding='utf-8') as f:
    f.write(j)