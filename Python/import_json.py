import json
import os
from openpyxl import load_workbook

wb = load_workbook('F:/Giorgio/Multimedia/Download/data.xlsx')
ws = wb.active

num_cells_x = int((ws.max_column - 1) / 3)
num_cells_y = int((ws.max_row - 1) / 6)

cards = []
for x in range(num_cells_x):
    for y in range(num_cells_y):
        card_cell = ws.cell(y * 6 + 2, x * 3 + 2)
        card = {}
        for i, char in enumerate(['P', 'O', 'A', '?', 'S']):
            card = {
                'cat': char,
                'num': 1 if char == 'S' or card_cell.offset(i + 1, 1).value == 1 else 0,
                'par': card_cell.offset(i + 1, 2).value,
            }
            if card_cell.offset(i + 1, 2).value != None:
                cards.append(card)

j = json.dumps(cards)
print(j)

with open(os.path.join(os.path.dirname(os.path.realpath(__file__)), 'cards.json'), 'w') as f:
    f.write(j)